import streamlit as st
import pandas as pd
import io
import re
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment

from calculations.arudha_calc import calc_all_arudhas
from calculations.ul_calc import calc_UL
from data.houses import generate_house_lords, generate_house_signs

from dictionaries import (
    AL, A2, A3, A4, A5, A6, A7,
    A8, A9, A10, A11, A12, UL
)


# --------------------------------------------
# 페이지 상태 초기화
# --------------------------------------------
if "page" not in st.session_state:
    st.session_state.page = "main"

if "results" not in st.session_state:
    st.session_state.results = None



# ============================================
# 메인 화면
# ============================================
def show_main_screen():

    st.title("☀️ Arudha & Upapada Calculator")

    left, right = st.columns([1, 1])

    with left:

        st.subheader("Input")

        asc = st.selectbox(
            "Ascendant Sign",
            ["Aries", "Taurus", "Gemini", "Cancer", "Leo", "Virgo",
             "Libra", "Scorpio", "Sagittarius", "Capricorn", "Aquarius", "Pisces"]
        )

        lord_positions = {}
        planets = ["Sun", "Moon", "Mercury", "Venus", "Mars", "Jupiter", "Saturn"]

        for planet in planets:
            lord_positions[planet] = st.selectbox(f"{planet} House", list(range(1, 13)))

        if st.button("Calculate Arudhas"):
            house_lords = generate_house_lords(asc)
            arudhas = calc_all_arudhas(lord_positions, house_lords)
            ul = calc_UL(lord_positions, house_lords)

            st.session_state.results = {
                "asc": asc,
                "lord_positions": lord_positions,
                "house_lords": house_lords,
                "arudhas": arudhas,
                "ul": ul
            }

            st.session_state.page = "main_results"
            st.rerun()

    with right:

        st.subheader("Results")

        if st.session_state.results:
            r = st.session_state.results
            arudhas = r["arudhas"]

            house_signs = generate_house_signs(r["asc"])

            for key, val in arudhas.items():
                rashi = house_signs[val]
                st.write(f"{key} → {rashi} (H{val})")

            ul_house = r["ul"]
            ul_rashi = house_signs[ul_house]
            st.write(f"UL → {ul_rashi} (H{ul_house})")

            if st.button("See Interpretation"):
                st.session_state.page = "interpret"
                st.rerun()


# ============================================
# ❗ 해석 화면 (함수 밖으로 꺼냄!)
# ============================================
def show_interpretation_screen():

    st.title("🔱 Arudha & UL Interpretation")

    results = st.session_state.results
    if not results:
        st.error("No results found.")
        return

    # dict 매핑
    arudha_maps = {
        "AL": AL.Arudha_dict,
        "A2": A2.Arudha_dict,
        "A3": A3.Arudha_dict,
        "A4": A4.Arudha_dict,
        "A5": A5.Arudha_dict,
        "A6": A6.Arudha_dict,
        "A7": A7.Arudha_dict,
        "A8": A8.Arudha_dict,
        "A9": A9.Arudha_dict,
        "A10": A10.Arudha_dict,
        "A11": A11.Arudha_dict,
        "A12": A12.Arudha_dict,
        "UL": UL.Arudha_dict
    }

    rows = []
    all_arudhas = results["arudhas"].copy()
    all_arudhas["UL"] = results["ul"]

    # Asc 기반 house→rashi
    house_signs = generate_house_signs(results["asc"])

    for key, house_num in all_arudhas.items():
        rashi = house_signs[house_num]
        dic = arudha_maps[key]
        house_text = dic["house"][house_num]
        rashi_text = dic["rashi"][rashi]

        explanation = house_text + "<br>" + rashi_text
        rows.append([key, house_num, rashi, explanation])

    df = pd.DataFrame(rows, columns=["Arudha", "House", "Rashi", "Interpretation"])

    # ============================
    # HTML 테이블 정상 출력
    # ============================
    df_display = df.copy()
    df_display["Interpretation"] = df_display["Interpretation"].str.replace("\n", "<br>")

    html_table = df_display.to_html(escape=False, index=False)

    st.markdown("""
    <style>
    table th { text-align:left !important; }
    table td { text-align:left !important; vertical-align:top !important; }
    </style>
    """, unsafe_allow_html=True)

    st.markdown(html_table, unsafe_allow_html=True)


    # ============================
    # Excel 파일 생성 (스타일링)
    # ============================
    def clean_for_excel(text):
        text = text.replace("<br>", "\n")
        text = re.sub(r"</?b>", "", text)
        return text

    excel_df = df.copy()
    excel_df["Interpretation"] = excel_df["Interpretation"].apply(clean_for_excel)

    wb = Workbook()
    ws = wb.active
    ws.title = "Arudhas"

    # 헤더
    headers = ["Arudha", "House", "Rashi", "Interpretation"]
    ws.append(headers)

    for cell in ws[1]:
        cell.font = Font(bold=True)

    # 데이터 입력
    for _, row in excel_df.iterrows():
        ws.append(row.tolist())

    # Interpretation 줄바꿈 & 스타일
    for row_idx in range(2, ws.max_row + 1):
        cell = ws[f"D{row_idx}"]
        cell.alignment = Alignment(wrap_text=True)

    # 열 너비 자동 조정
    for col in ws.columns:
        max_len = max(len(str(cell.value)) if cell.value else 0 for cell in col)
        ws.column_dimensions[col[0].column_letter].width = min(max_len * 1.1, 60)

    # 파일 저장
    output = io.BytesIO()
    wb.save(output)
    xlsx_data = output.getvalue()

    st.download_button(
        label="Download Excel File",
        data=xlsx_data,
        file_name="arudha_interpretation.xlsx",
        mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
    )


# ============================================
# 페이지 라우팅
# ============================================
if st.session_state.page == "main":
    show_main_screen()

elif st.session_state.page == "main_results":
    show_main_screen()

elif st.session_state.page == "interpret":
    show_interpretation_screen()



